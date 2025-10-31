# services/rag.py
from __future__ import annotations
import os, glob, csv, re, sys
from typing import Optional, List, Tuple, Dict
from datetime import datetime, timedelta

import pandas as pd
from openpyxl import load_workbook
import httpx
from chromadb import PersistentClient
from chromadb.utils import embedding_functions
from dotenv import load_dotenv

load_dotenv()

print("🧠 Python 실행 경로:", sys.executable)
print("📦 httpx 버전:", httpx.__version__)

# ---------------- OpenAI 클라이언트 ----------------
try:
    from openai import OpenAI as _OpenAI
except Exception:
    _OpenAI = None

def _make_openai_client() -> Optional["_OpenAI"]:
    if not _OpenAI:
        return None
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        print("❌ OPENAI_API_KEY가 .env에서 로드되지 않았습니다.")
        return None
    os.environ["OPENAI_API_KEY"] = api_key

    http_proxy  = os.getenv("HTTP_PROXY")  or os.getenv("http_proxy")
    https_proxy = os.getenv("HTTPS_PROXY") or os.getenv("https_proxy")

    client_args = {"timeout": 30.0}
    if http_proxy or https_proxy:
        proxies = {}
        if http_proxy:
            proxies["http://"] = http_proxy
        if https_proxy:
            proxies["https://"] = https_proxy
        client_args["proxies"] = proxies

    http_client = httpx.Client(**client_args) if client_args else None
    return _OpenAI(http_client=http_client)

_OPENAI_CLIENT = _make_openai_client()
_OPENAI_EMB_MODEL = os.getenv("EMB_MODEL", "text-embedding-3-small")

# ---------------- 경로 설정 ----------------
DBDIR  = "data/vectorstore/chroma"
FAQDIR = "data/vectorstore"

# ---------------- Chroma Client ----------------
def _chroma_client():
    os.makedirs(DBDIR, exist_ok=True)
    return PersistentClient(path=DBDIR)

# ---------------- OpenAI 임베딩 ----------------
class OpenAIEmbedder:
    """ChromaDB 0.4.16+ 호환 임베더 (긴 입력 자동 분할)"""
    def __init__(self, client, model_name: str):
        self.client = client
        self.model_name = model_name
        self._name = f"openai:{model_name}"

    def name(self):
        """Chroma가 호출할 때 사용하는 식별자"""
        return self._name

    def __call__(self, input: list[str]) -> list[list[float]]:
        """
        긴 입력 자동 분할 (8192 tokens 초과 방지)
        - 한글 기준 6000자 단위로 chunk
        - 실패한 chunk는 dummy 벡터로 채워 1:1 매핑 유지
        """
        def chunk_text(text: str, max_chars: int = 6000) -> list[str]:
            return [text[i:i + max_chars] for i in range(0, len(text), max_chars)]

        safe_inputs = []
        for t in input:
            if isinstance(t, str) and len(t) > 6000:
                safe_inputs.extend(chunk_text(t))
            else:
                safe_inputs.append(t)

        all_embeddings = []
        for chunk in safe_inputs:
            try:
                resp = self.client.embeddings.create(model=self.model_name, input=chunk)
                all_embeddings.extend([d.embedding for d in resp.data])
            except Exception as e:
                print(f"❌ Embedding error for chunk ({len(str(chunk))} chars): {e}")
                dim = 1536 if "3-small" in self.model_name else 3072
                all_embeddings.append([0.0] * dim)

        # 개수 보정
        if len(all_embeddings) < len(safe_inputs):
            diff = len(safe_inputs) - len(all_embeddings)
            dim = 1536 if "3-small" in self.model_name else 3072
            for _ in range(diff):
                all_embeddings.append([0.0] * dim)
        elif len(all_embeddings) > len(safe_inputs):
            all_embeddings = all_embeddings[:len(safe_inputs)]

        return all_embeddings

    # ✅ Chroma가 query 시 호출하는 메서드
    def embed_query(self, input: list[str]) -> list[list[float]]:
        return self.__call__(input)

    # ✅ Chroma가 upsert 시 호출하는 메서드
    def embed_documents(self, input: list[str]) -> list[list[float]]:
        return self.__call__(input)


def _embedding_function():
    if _OPENAI_CLIENT:
        return OpenAIEmbedder(_OPENAI_CLIENT, _OPENAI_EMB_MODEL)
    return embedding_functions.SentenceTransformerEmbeddingFunction(
        model_name="sentence-transformers/all-MiniLM-L6-v2"
    )

# ---------------- 파일 리더 ----------------
def _read(path: str) -> str:
    p = path.lower()
    if p.endswith((".md", ".txt")):
        return open(path, "r", encoding="utf-8").read()
    if p.endswith(".pdf"):
        try:
            from pypdf import PdfReader
            return "\n".join(page.extract_text() or "" for page in PdfReader(path).pages)
        except Exception:
            return ""
    return ""

# ---------------- QA 포맷 ----------------
def _read_qa_csv(path: str) -> list[tuple[str, str]]:
    rows = []
    with open(path, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            q, a = (r.get("question") or "").strip(), (r.get("answer") or "").strip()
            if q and a:
                rows.append((q, a))
    return rows

def _read_qa_md(path: str) -> list[tuple[str, str]]:
    txt = open(path, "r", encoding="utf-8").read()
    pairs = []
    for block in txt.split("### Q:")[1:]:
        q, _, rest = block.partition("\n")
        a_tag = "A:"
        a = rest.split("\n", 1)[0] if a_tag not in rest else rest.split(a_tag, 1)[1].strip()
        q, a = q.strip(), a.strip()
        if q and a:
            pairs.append((q, a))
    return pairs

# ---------------- 숨김 / 정정 ----------------
def _read_hidden_data():
    hidden_dir = os.path.join(FAQDIR, "hidden")
    correction_dir = os.path.join(FAQDIR, "correction")
    hidden_txt, correction_txt = [], []
    for folder, acc in [(hidden_dir, hidden_txt), (correction_dir, correction_txt)]:
        if not os.path.exists(folder):
            continue
        for f in glob.glob(os.path.join(folder, "*.txt")) + glob.glob(os.path.join(folder, "*.csv")):
            try:
                acc.append(open(f, encoding="utf-8").read())
            except Exception as e:
                print(f"⚠ {folder} 파일 읽기 오류: {e}")
    return "\n".join(hidden_txt), "\n".join(correction_txt)

# ---------------- 조직도 ----------------
def _read_org_info() -> Optional[pd.DataFrame]:
    org_path = os.path.join(FAQDIR, "org_info.csv")
    if not os.path.exists(org_path):
        print("⚠️ 조직도 파일(org_info.csv)이 없습니다.")
        return None
    try:
        return pd.read_csv(org_path, encoding="utf-8")
    except Exception as e:
        print(f"❌ 조직도 CSV 파싱 오류: {e}")
        return None

def _count_members_from_org(org_path: str, leader_name: str) -> tuple[int | None, str | None]:
    try:
        df = pd.read_csv(org_path)
        row = df[df["팀장"] == leader_name]
        if row.empty:
            return None, None
        if "팀원수" in df.columns:
            team_count = int(row.iloc[0]["팀원수"])
        else:
            members_str = str(row.iloc[0].get("팀원", "")).strip()
            team_count = len([m for m in members_str.split(",") if m.strip()])
        return team_count, row.iloc[0].get("팀명", None)
    except Exception as e:
        print(f"❌ 조직도 CSV 파싱 오류: {e}")
        return None, None

# ---------------- 색인 빌드 ----------------
def build_index_all(verbose=True):
    cli = _chroma_client()
    try:
        cli.delete_collection("faqs_openai")
        if verbose: print("✅ 기존 컬렉션 삭제 완료")
    except Exception:
        if verbose: print("ℹ️ 기존 컬렉션 없음")

    col = cli.get_or_create_collection("faqs_openai", embedding_function=_embedding_function())
    ids, docs, metas = [], [], []
    exts = [".md", ".txt", ".pdf", ".csv", ".xlsx"]
    files = sum([glob.glob(os.path.join(FAQDIR, f"*{ext}")) for ext in exts], [])
    MAX_CHARS = 6000

    for f in files:
        text = ""
        try:
            if f.endswith((".md", ".txt")):
                text = open(f, "r", encoding="utf-8").read()
            elif f.endswith(".pdf"):
                from pypdf import PdfReader
                text = "\n".join([p.extract_text() or "" for p in PdfReader(f).pages])
            elif f.endswith(".csv"):
                df = pd.read_csv(f, encoding="utf-8")
                text = "\n".join([", ".join([f"{c}: {v}" for c, v in row.items()]) for _, row in df.iterrows()])
            elif f.endswith(".xlsx"):
                wb = load_workbook(f, data_only=True)
                for s in wb.sheetnames:
                    ws = wb[s]
                    for row in ws.iter_rows(values_only=True):
                        text += ", ".join([str(v) for v in row if v]) + "\n"
        except Exception as e:
            print(f"⚠️ {os.path.basename(f)} 읽기 오류: {e}")
            continue

        if text.strip():
            for i in range(0, len(text), MAX_CHARS):
                chunk = text[i:i+MAX_CHARS]
                ids.append(f"{os.path.basename(f)}_part{i//MAX_CHARS+1}")
                docs.append(chunk)
                metas.append({"path": f, "part": i//MAX_CHARS+1})

    if ids:
        col.upsert(ids=ids, documents=docs, metadatas=metas)
        if verbose: print(f"✅ {len(ids)}개 문서 색인 완료")

    qa_files = glob.glob(f"{FAQDIR}/*.csv") + glob.glob(f"{FAQDIR}/faq_qa.md") + glob.glob(f"{FAQDIR}/*_qa.md")
    q_ids, q_docs, q_metas = [], [], []
    for f in qa_files:
        pairs = _read_qa_csv(f) if f.endswith(".csv") else _read_qa_md(f)
        for i, (q, a) in enumerate(pairs):
            q_ids.append(f"{os.path.basename(f)}::Q{i+1}")
            q_docs.append(f"[Q] {q}\n[A] {a}")
            q_metas.append({"path": f, "type": "qa"})
    if q_ids:
        col.upsert(ids=q_ids, documents=q_docs, metadatas=q_metas)
        if verbose: print(f"✅ {len(q_ids)}개 Q&A 색인 완료")

    org_df = _read_org_info()
    if org_df is not None:
        org_text = "\n".join([
            f"{r['팀명']}팀 ({r['팀장']}): {r['팀원수']}명, {r['팀원']}"
            for _, r in org_df.iterrows()
        ])
        col.upsert(ids=["org_info"], documents=[org_text], metadatas=[{"path": "org_info.csv", "type": "org"}])
        if verbose: print("✅ 조직도 색인 완료")


# ---------------- RAG 질의 ----------------
def safe_rag_query(question: str, k=3, show_sources=False):
    cli = _chroma_client()
    col = cli.get_or_create_collection("faqs_openai", embedding_function=_embedding_function())
    r = col.query(query_texts=[question], n_results=k)

    ctx = "\n\n".join(r["documents"][0]) if r.get("documents") else ""
    metas = r["metadatas"][0] if r.get("metadatas") else []

    now = datetime.now()
    today_str = now.strftime("%Y-%m-%d (%A)")
    week_start = (now - timedelta(days=now.weekday())).strftime("%Y-%m-%d")
    week_end = (now + timedelta(days=6 - now.weekday())).strftime("%Y-%m-%d")
    month_str = now.strftime("%Y-%m")

    hidden_context, correction_rules = _read_hidden_data()

    designated_amounts = {}
    for line in hidden_context.splitlines():
        m = re.match(r"(\S+)\s*(\d+(?:,\d+)*)원", line)
        if m:
            rank, amt = m.groups()
            designated_amounts[rank] = int(amt.replace(",", ""))

    match_name_rank = re.search(r"([가-힣A-Za-z]+)\s*(팀장|부장|임원|차장|과장|대리|사원)", question)
    user_name, user_rank = (match_name_rank.groups() if match_name_rank else (None, None))

    org_path = os.path.join(FAQDIR, "org_info.csv")
    member_count, team_name = _count_members_from_org(org_path, user_name) if user_name else (None, None)

    team_info_text, auto_calc_text = "", ""
    if member_count and team_name:
        team_info_text = f"{team_name} ({user_rank} {user_name})의 팀원 수는 {member_count}명입니다."
    if "업무추진비" in question and user_rank in designated_amounts:
        base = designated_amounts[user_rank]
        if member_count:
            total = base * member_count
            auto_calc_text = f"💰 자동 계산: {user_rank} {user_name} - 팀원 {member_count}명 × {base:,}원 = {total:,}원"
        else:
            auto_calc_text = f"💰 자동 계산: {user_rank} 기준 1인당 {base:,}원"

    system_prompt = (
        f"오늘은 {today_str}이며, 이번 주는 {week_start}~{week_end}, 이번 달은 {month_str}월입니다.\n"
        "다음 정보를 참고하여 간결하고 정확하게 한국어로 답변하세요.\n"
        "hidden 폴더의 내용은 내부 규칙 참고용이며, 답변에 직접 노출하지 마세요.\n\n"
        f"[숨김 규칙]\n{hidden_context}\n"
        f"[정정 자료]\n{correction_rules}\n"
        f"[조직도 정보]\n{team_info_text}\n"
        f"[자동 계산]\n{auto_calc_text}\n"
    )

    if _OPENAI_CLIENT and ctx.strip():
        try:
            resp = _OPENAI_CLIENT.chat.completions.create(
                model=os.getenv("OPENAI_MODEL", "gpt-4o-mini"),
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": f"[컨텍스트]\n{ctx}\n\n[질문]\n{question}"},
                ],
                temperature=0.1,
            )
            answer = resp.choices[0].message.content
        except Exception as e:
            answer = f"❌ LLM 오류: {e}"
    else:
        answer = f"(OPENAI_API_KEY 없음 또는 컨텍스트 부족)\n\n{ctx[:800]}"

    if show_sources and metas:
        srcs = [f"📄 `{os.path.basename(m.get('path',''))}`" for m in metas if m.get("path")]
        if srcs:
            answer += "\n\n---\n**📂 참고 문서:**\n" + "\n".join(srcs)
    return answer

# rag.py 파일 끝에 추가

# ---------------- 법인카드 매칭 기능 (수정) ----------------
import re

def match_corporate_card(approval_df: pd.DataFrame, expense_df: pd.DataFrame, limits_df: pd.DataFrame) -> tuple:
    """법인카드 승인내역과 지출결의 매칭"""
    
    # 한도금액, 승인번호 열 추가
    expense_df.insert(0, '한도금액', '')
    expense_df['승인번호'] = ''
    
    # 승인내역에서 매칭된 행 인덱스 저장
    matched_approval_indices = []
    
    # 각 지출결의 행 처리
    for idx, exp_row in expense_df.iterrows():
        basic_summary = str(exp_row.get('기본적요', ''))
        
        # 기본적요에서 첫 4자리 숫자 추출 (지출결의)
        four_digit_match = re.match(r'^(\d{4})', basic_summary)
        if four_digit_match:
            four_digit_key = four_digit_match.group(1)
            
            # limits.csv에서 매칭하여 한도금액 설정
            for _, limit_row in limits_df.iterrows():
                if str(limit_row['적요']) == four_digit_key:
                    amount = int(limit_row['금액'])
                    # -1, -2 같은 특수값 처리
                    if amount > 0:
                        expense_df.at[idx, '한도금액'] = f"{amount:,}"
                    elif amount == -1:
                        expense_df.at[idx, '한도금액'] = "한도없음"
                    elif amount == -2:
                        expense_df.at[idx, '한도금액'] = "실비정산"
                    else:
                        expense_df.at[idx, '한도금액'] = str(amount)
                    break
        
        # '법카'가 없고 '개인' 포함시 승인번호 매칭 스킵
        if '법카' not in basic_summary and '개인' in basic_summary:
            continue
        
        # 법인카드인 경우 승인내역과 매칭
        if '법카' in basic_summary:
            exp_time = str(exp_row.get('승인시간', ''))
            
            # 승인내역에서 동일 시간 찾기
            for app_idx, app_row in approval_df.iterrows():
                if str(app_row.get('승인시간', '')) == exp_time and exp_time != '':
                    # 날짜 형식 변환 및 비교
                    exp_date = str(exp_row.get('증빙일자', ''))
                    # "2025-09-04(목)" 형식에서 날짜만 추출
                    if '(' in exp_date:
                        exp_date = exp_date.split('(')[0]
                    exp_date = exp_date.replace('-', '.')
                    
                    app_date = str(app_row.get('승인일', ''))
                    if exp_date == app_date:
                        # 승인번호 매칭
                        expense_df.at[idx, '승인번호'] = str(app_row.get('승인번호', ''))
                        matched_approval_indices.append(app_idx)
                        break
    
    return expense_df, matched_approval_indices

def extract_four_digit_code(text: str) -> str:
    """기본적요에서 처음 네 자리 숫자 추출"""
    match = re.match(r'(\d{4})', str(text))
    return match.group(1) if match else ""

# ---------------- 엔터키 입력 헬퍼 ----------------
def format_question_with_enter(question: str) -> str:
    """엔터키 입력 지원을 위한 질문 포맷팅"""
    return question.strip().replace('\n', ' ')